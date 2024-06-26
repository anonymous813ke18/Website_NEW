from flask import Flask, request, render_template, jsonify
import subprocess
import json
import pandas as pd
from flask_cors import CORS  # Import CORS
from flask_cors import cross_origin
from itertools import count
import re
import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask, request, redirect, url_for
from bs4 import BeautifulSoup
from flask import session
from datetime import date
from dateutil import tz
from datetime import datetime as dt, timezone
import datetime
import datetime
from datetime import datetime
from flask import Flask, jsonify
import pandas as pd
from flask import Flask, request, render_template, jsonify
from flask import Flask, request, jsonify
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from flask import Flask, request, jsonify
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, jsonify, request
import pandas as pd
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from flask import Flask, request, render_template_string
import pandas as pd
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from flask import request


app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
cors = CORS(app, resources={r"/handover_form": {"origins": "*"}})  # Enable CORS for /handover_form route
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'

@app.route('/')
def index():
    return render_template('Modules/Main/Start/templates/login.html')

@app.route('/send_items')
def send_items():
    return render_template('Modules/Main/Handover/New UI/templates/handover.html')

@app.route('/welcome')
def welcome():
    return render_template('Modules/Main/Start/templates/welcome.html')

@app.route('/return_to_login')
def return_to_login():
    return render_template('Modules/Main/Start/templates/login.html')

@app.route('/manager')
def manager():
    return render_template('Modules/Main/Start/templates/manager.html')

@app.route('/employee')
def employee():
    return render_template('Modules/Main/Start/templates/employee.html')

@app.route('/approve_receive')
def approve_receive():
    return render_template('Modules/Main/Approval/Approval Receive/templates/approve_receive_bar.html')

@app.route('/approve_send')
def approve_send():
    return render_template('Modules/Main/Approval/Approval Send/templates/approve_send_bar.html')

@app.route('/transactionhistory')
def transactionhistory():
    return render_template('Modules/Main/Transaction History/transactionHistory.html')

@app.route('/transactionstatus')
def transactionstatus():
    return render_template('Modules/Main/Transaction Status/transactionstatus.html')

@app.route('/approvetable')
def approvetable():
    return render_template('Modules/Main/Approval/approvalTable.html')

@app.route('/approve_table')
def approve_table():
    return render_template('approvalTable.html')

@app.route('/invent')
def invent():
    return render_template('Modules/Main/Dashboard/templates/invent.html')

@app.route('/project_invent')
def project_invent():
    return render_template('Modules/Main/Dashboard/templates/project_invent.html')

@app.route('/my_invent')
def my_invent():
    return render_template('Modules/Main/Dashboard/templates/my_invent.html')

@app.route('/receive_form_data')
def receive_form_data():
    return render_template('Modules/Main/Receiver/form_data.html')

@app.route('/receive_table')
def receive_table():
    return render_template('Modules/Main/Receiver/recieveTable.html')

@app.route('/transaction_history_form_data')
def transaction_history_form_data():
    return render_template('Modules/Main/Transaction History/transaction_history_form_data.html')
    
@app.route('/display_send_approval')
def display_send_approval():
    return render_template('Modules/Main/Approval/approve_send_form_data.html')


@app.route('/display_receive_approval')
def display_receive_approval():
    return render_template('Modules/Main/Approval/approve_receive_form_data.html')

@app.route('/display_transaction_progess_table')
def display_transaction_progess_table():
    return render_template('Modules/Main/Transaction Status/transactionprogressform.html')



'''
def extract_data_from_excel(product_ids):
    # Load the Excel file
    df = pd.read_excel('Excel/handover_data.xlsx')
    
    # Initialize a list to store the extracted data
    extracted_data = []
    
    # Iterate over each product id
    for product_id in product_ids:
        # Filter rows where 'ProductID' column matches the product id
        filtered_rows = df[df['ProductID'] == product_id]
        
        # Iterate over each filtered row
        for index, row in filtered_rows.iterrows():
            # Extract the values from the row and append to the extracted_data list
            extracted_data.append([row['ProductID'], row['Name'], row['FormID'], row['HandoverDate']])
    
    return extracted_data



def extract_rows_first_filter(excel_file, column_name, value):
    try:
        df = pd.read_excel(excel_file)
    except FileNotFoundError:
        print("Excel file not found.")
        return None, None
    except Exception as e:
        print("An error occurred while loading the Excel file:", e)
        return None, None
    
    if column_name not in df.columns:
        print(f"Column '{column_name}' not found in the Excel file.")
        return None, None

    # Replace all NaN values in the DataFrame with 'NAN'
    df = df.fillna('nan')

    # Convert both the DataFrame values and the comparison value to lowercase
    df[column_name] = df[column_name].str.lower().str.strip()
    value = value.lower().strip()


    #print("this is the excel data", df)
    print("this column name", column_name)
    print("this is the value we are searching", value)

    filtered_df = df[df[column_name] == value]
    print("lets see this dataframe", filtered_df)
    # Count unique values in the 'FormID' column
    form_id_count = filtered_df['FormID'].nunique()
    
    return filtered_df, form_id_count

def correct_versions(versions):
    try:
        # Extracting major parts
        major_parts = [int(version.split('.')[0]) for version in versions]
        print("Extracted major parts:", major_parts)

        # Mapping major parts to consecutive digits starting from 1
        major_mapping = {}
        mapped_major_parts = []
        counter = 1
        for major in major_parts:
            if major not in major_mapping:
                major_mapping[major] = counter
                counter += 1
            mapped_major_parts.append(major_mapping[major])
        print("Mapped major parts:", mapped_major_parts)

        # Reconstructing the strings with mapped major parts
        mapped_versions = [f"{mapped_major}.{minor}" for mapped_major, minor in zip(mapped_major_parts, [version.split('.')[1] for version in versions])]
        print("Mapped versions:", mapped_versions)

        return mapped_versions
    except Exception as e:
        print("An error occurred in correcting versions:", e)
        return None
'''

filtered_df = None


def extract_rows_from_excel(form_id):

    global filtered_df


    # Read Excel file into a DataFrame
    df = pd.read_excel('Excel/handover_data.xlsx')  # Replace 'your_excel_file.xlsx' with your file path

    print("this is the form ID we are trying to find in the df", form_id)

    # Filter rows based on form ID
    formid_filtered_df = df[df["FormID"] == form_id]

    if formid_filtered_df.empty:
        print("Form ID not found in the DataFrame")
        return []

    # Get the main digit from the first row where form ID matches
    main_digit = int(str(formid_filtered_df["Serial"].iloc[0]).split(".")[0])

    # Filter rows based on the main digit
    filtered_df = df[df["Serial"].astype(str).str.split(".").str[0] == str(main_digit)]

    # Replace NaN values with "Nan"
    filtered_df.fillna("Nan", inplace=True)

    print("This is the filtered df", filtered_df)
    return filtered_df.to_dict(orient="records")



@app.route('/send_formid')
def send_formid():
    form_id = request.args.get('form_id')
    print("Received Form ID:", form_id)
    extract_rows_from_excel(form_id)
    # Do whatever you need to do with the form ID
    return "Form ID received successfully"


@app.route('/get_form_data')
def get_form_data():
    global filtered_df
    
    if filtered_df is None:
        print("Filtered DataFrame is not available")
        return jsonify([])

    # Convert DataFrame to JSON string and return
    json_data = filtered_df.to_dict(orient="records")
    return jsonify(json_data)


@app.route('/login', methods=['POST'])
def login():

    # Load the user_data Excel file
    user_data = pd.read_excel("Excel/login_details.xlsx")


    user_id = request.form['id']
    password = request.form['password']
    print(user_id,password)

    # Check if ID and password exist in the user_data DataFrame
    matched_rows = user_data[(user_data['ID'] == user_id) & (user_data['Password'] == password)]

    if not matched_rows.empty:
        # Extract type of account
        account_type = matched_rows.iloc[0]['Type Of Account']
        
        # Store the entire row data in session
        session['login_row_data'] = matched_rows.iloc[0].to_dict()
        print("this is session dataaa  ", session['login_row_data'])

        # Redirect based on account type
        if account_type == 'Manager':
            return redirect(url_for('manager'))

        elif account_type == 'Employee':
            return redirect(url_for('employee'))

        else:
            return "Unknown account type"

    else:
        return "Account not found"




global from_person 
global to_person 
global from_project 
global to_project 


@app.route('/send_approval_request', methods=['POST'])
def send_approval_request():
    
    global from_person 
    global to_person 
    global from_project 
    global to_project 

    form_data = request.json  # Get JSON data from the request
    print("This is the form data for approval request", form_data)

    # Check if form_data is not empty and is a list
    if form_data and isinstance(form_data, list) and len(form_data) > 0:
        first_dict = form_data[0]  # Get the first dictionary from the list

        # Access and print values of 'FromPerson', 'ToPerson', 'FromProject', and 'ToProject'
        from_person = first_dict.get('FromPerson')
        to_person = first_dict.get('ToPerson')
        from_project = first_dict.get('FromProject')
        to_project = first_dict.get('ToProject')

        print("From Person:", from_person)
        print("To Person:", to_person)
        print("From Project:", from_project)
        print("To Project:", to_project)
    else:
        print("No form data or invalid data format received")


    # Mapping of keys to column indices
    key_to_column = {
        'FormID': 0,  # Add FormID to the mapping
        'Category': 2,
        'ProductID': 3,
        'Name': 4,
        'Make': 5,
        'Model': 6,
        'SenderCondition': 11,
        'SenderRemarks': 12,
    }

    # Define column names
    columns = ['FormID', 'Serial', 'Category', 'ProductID', 'Name', 'Make', 'Model', 
                'FromProject', 'ToProject','FromPerson', 'ToPerson',  
                'SenderCondition', 'SenderRemarks', 'ReceiverCondition', 'ReceiverRemarks' ]  # Add FormID to the columns list

    # Load existing Excel file if it exists, otherwise create an empty DataFrame
    try:
        existing_df = pd.read_excel('Excel/handover_data.xlsx')
    except FileNotFoundError:
        existing_df = pd.DataFrame(columns=columns)
    except ValueError:
        existing_df = pd.DataFrame(columns=columns)

    # Create a DataFrame from the form data
    new_data = []
    for item in form_data[1:]:  # Exclude the first row which contains metadata
        new_row = [None] * len(columns)  # Initialize a new row with None values
        for key, value in item.items():
            if key in key_to_column:
                if isinstance(key_to_column[key], int):
                    new_row[key_to_column[key]] = value
                elif isinstance(key_to_column[key], dict):
                    # Set flags for 'frontend', 'backend', and 'cleanup' based on the 'Status' field
                    status_parts = value.split(',')
                    for status_part in status_parts:
                        if status_part.strip() in key_to_column[key]:
                            new_row[key_to_column[key][status_part.strip()]] = 1
        new_data.append(new_row)

    new_df = pd.DataFrame(new_data, columns=columns)

    # Generate unique FormID
    def generate_form_id():
        original_id = 'abcd1234'
        id_list = list(original_id)
        random.shuffle(id_list)
        return ''.join(id_list)

    # Extract values from the "FormID" column
    form_ids = existing_df['FormID'].tolist()

    # Generate a unique FormID
    unique_form_id = generate_form_id()
    while unique_form_id in form_ids:
        unique_form_id = generate_form_id()

    # Insert the unique FormID into the "FormID" column of the first product
    if len(new_df) > 0:
        new_df.loc[0, 'FormID'] = unique_form_id

    if len(new_df) > 0:
        new_df.loc[0, 'FromProject'] = from_project
    
    if len(new_df) > 0:
        new_df.loc[0, 'ToProject'] = to_project
    
    if len(new_df) > 0:
        new_df.loc[0, 'FromPerson'] = from_person

    if len(new_df) > 0:
        new_df.loc[0, 'ToPerson'] = to_person

    text="Send Form"
    # Call the function to send email with the generated PDF attached
    send_email(text, unique_form_id ,None,from_person,to_person,from_project ,to_project )


    # Determine the next value for "Serial No" for the new data
    if not existing_df.empty and 'Serial' in existing_df:
        last_serial_no = existing_df['Serial'].iloc[-1]  # Get the last serial number
        last_serial_no = str(last_serial_no)  # Convert to string explicitly
        match = re.match(r"(\d+)\.(\d+)", last_serial_no)  # Extract numerical parts using regex
        if match:
            last_main_num = int(match.group(1))  # Extract the main number part
            last_sub_num = int(match.group(2))  # Extract the sub number part
        else:
            last_main_num = 0
            last_sub_num = 0
    else:
        last_main_num = 0
        last_sub_num = 0

    # Calculate the total number of rows in the existing DataFrame and new DataFrame combined
    total_rows = len(existing_df) + len(new_df)

    # Determine the current numerical part for the new batch
    current_main_num = last_main_num + 1

    # Reset the subpart if the main part changes or if there are no existing serial numbers
    if current_main_num > last_main_num or total_rows == 0:
        last_sub_num = 0

    # Iterate over each row of new data and generate serial numbers
    serial_numbers = []
    for i in range(len(new_df)):
        current_sub_num = last_sub_num + i + 1
        new_serial_no = str(current_main_num) + '.' + str(current_sub_num)
        serial_numbers.append(new_serial_no)

    # Add serial numbers to the new data frame
    new_df['Serial'] = serial_numbers

    # Concatenate existing and new data frames
    df = pd.concat([existing_df, new_df], ignore_index=True)

    # Write the updated data frame back to the Excel file
    with pd.ExcelWriter('Excel/handover_data.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Return a success message
    return jsonify({'message': 'Excel file updated successfully'})


def send_email(message_content, form_no, eway_bill_no=None,from_person=None,to_person=None,from_project =None ,to_project =None ):

    print("This is the email From Person:", from_person)
    print("This is the email To Person:", to_person)
    print("This is the email From Project:", from_project)
    print("This is the email To Project:", to_project)

    # Sender and receiver email addresses
    sender_email = "shaikhfahad687@gmail.com"  # Update with your Gmail address
    receiver_email = "shaikhfahad687@gmail.com"  # Update with the receiver's email address

    try:
        # Gmail SMTP server details
        smtp_server = "smtp.gmail.com"
        smtp_port = 587  # Port for TLS encryption

        # Login credentials (use app password)
        email_password = "rgbbdlhpyleheico"  # Update with your Gmail app password

        # Establish a secure connection with the SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            # Login to the SMTP server
            server.login(sender_email, email_password)
            
            # Create a multipart message
            message = MIMEMultipart()
            message["From"] = sender_email
            message["To"] = receiver_email
            
            if message_content == "Send Form":
                message["Subject"] = "Handover Transaction details"
                body_message = f"I want to send items, The excel sheet of the eway bill is attached below. Please generate the eway bill and approve this form. The form details are as follows:\n\nForm no: {form_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
                # Get the current working directory
                current_directory = os.getcwd()
                # Assuming the Excel sheet is in the same directory, change the filename if it's different
                excel_file_path = os.path.join(current_directory, "Excel/eway_bill.xlsx")

                # Open the Excel file in binary mode
                with open(excel_file_path, "rb") as attachment:
                    # Add Excel file as application/octet-stream
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                # Encode file in ASCII characters to send by email    
                encoders.encode_base64(part)

                # Add header as key/value pair to attachment part
                part.add_header("Content-Disposition", f"attachment; filename= {os.path.basename(excel_file_path)}")

                # Add attachment to message
                message.attach(part)
                
            elif message_content == "Receive Form":
                message["Subject"] = "Receive Form details"
                body_message = f"I have received the items, need your approval to proceed with using them, please approve this form. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"
                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
            elif message_content == "Send Approval Form":
                message["Subject"] = "Send Approval Form details"
                body_message = f"I have granted permission to transfer items. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"
                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
            elif message_content == "Receive Approval Form":
                message["Subject"] = "Receive Approval Form details"
                body_message = f"I have granted permission to accept the items received. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"

                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))

            # Convert the message to string
            str_message = message.as_string()
            
            # Send email
            server.sendmail(sender_email, receiver_email, str_message)
    except Exception as e:
        print(f"An error occurred: {e}")


@app.route('/receive_approval_request', methods=['POST'])
def receive_approval_request():

    form_data = request.json  # Assuming the form data is sent as JSON
    print("This is the receive_approval_request form data", form_data)

    if form_data:
        last_dict = form_data[0]  # Get the last dictionary from the list

        # Assign the values of the last dictionary to the global variables
        from_person = last_dict.get('FromPerson')
        to_person = last_dict.get('ToPerson')
        from_project = last_dict.get('FromProject')
        to_project = last_dict.get('ToProject')

        # Optionally, you can print the values for verification
        print("From Person:", from_person)
        print("To Person:", to_person)
        print("From Project:", from_project)
        print("To Project:", to_project)

    # Open handover_data.xlsx
    excel_file = "Excel/handover_data.xlsx"
    df = pd.read_excel(excel_file)

    # Get the EwayBill value from the first dictionary
    eway_bill_value = form_data[1].get('EwayBill', None)
    formid = form_data[2].get('FormNo', None)

    print("formid",formid)
    print("ewaybill no",eway_bill_value)
    
    text = "Receive Form"

    send_email(text, formid, eway_bill_value,from_person,to_person,from_project ,to_project)

    
    # Check if there is any form data
    if form_data:
        for item in form_data[3:]:  # Start iterating from the third dictionary onwards
            product_id = item['ProductID']
            receiver_condition = item['ReceiverCondition']
            receiver_remarks = item['ReceiverRemarks']

            # Find the row with the matching ProductID and update ReceiverCondition and ReceiverRemarks
            df.loc[df['ProductID'] == product_id, 'ReceiverCondition'] = receiver_condition
            df.loc[df['ProductID'] == product_id, 'ReceiverRemarks'] = receiver_remarks

        # Update the 'AskReceiveApproval' column to 'yes' for the first product
        first_product_id = form_data[3]['ProductID']  # Get the product ID from the third dictionary
        df.loc[df['ProductID'] == first_product_id, 'AskReceiveApproval'] = 'yes'

        # Save the updated DataFrame back to the Excel file
        df.to_excel(excel_file, index=False)

        return "Mail for approval of receiving item is sent, you may contact your manager to approve it."
    else:
        return "No data provided."

@app.route('/approve_send_request', methods=['POST'])
def approve_send_request():
    form_data = request.json  # Assuming the form data is sent as JSON
    print("This is the approve_send_request form data", form_data)

    if form_data:
        last_dict = form_data[-1]  # Get the last dictionary from the list

        # Assign the values of the last dictionary to the global variables
        from_person = last_dict.get('FromPerson')
        to_person = last_dict.get('ToPerson')
        from_project = last_dict.get('FromProject')
        to_project = last_dict.get('ToProject')

        # Optionally, you can print the values for verification
        print("From Person:", from_person)
        print("To Person:", to_person)
        print("From Project:", from_project)
        print("To Project:", to_project)

    # Open handover_data.xlsx
    excel_file = "Excel/handover_data.xlsx"
    df = pd.read_excel(excel_file)

    # Check if there is any form data
    if form_data:
        # Get the EwayBill value and FormNo from the form data
        eway_bill_value = form_data[0].get('EwayBill', None)
        formid = form_data[1].get('FormNo', None)
        print("formid", formid)
        print("This is the ewaybill noooo", eway_bill_value)
        


        # If EwayBill value is present, find the index of the row where 'FormNo' matches the received value
        if eway_bill_value:
            form_row_index = df.index[df['FormID'] == formid].tolist()
            if form_row_index:
                df.loc[form_row_index[0], 'EwayBillNo'] = eway_bill_value
        
        # Update the 'ApprovalToSend' column for the row where 'FormNo' matches the received value
        form_row_index = df.index[df['FormID'] == formid].tolist()
        if form_row_index:
            df.loc[form_row_index[0], 'ApprovalToSend'] = 'yes'
        
        # Save the updated DataFrame back to the Excel file
        df.to_excel(excel_file, index=False)

        text = "Send Approval Form"
        send_email(text, formid, eway_bill_value,from_person,to_person,from_project ,to_project)

        return "Approval has been successfully given, the email is sent, the sender may proceed to send the items"
    else:
        return "No data provided."


@app.route('/disapprove_send_request', methods=['POST'])
def disapprove_send_request():
    form_data = request.json  # Assuming the form data is sent as JSON
    print("This is the disapprove_send_request form data", form_data)
    
    # Read the Excel file into a DataFrame
    df = pd.read_excel('Excel/handover_data.xlsx')

    # Convert the list of dictionaries to a DataFrame
    form_df = pd.DataFrame(form_data)

    # Remove rows where 'ProductID' matches the one received in the form data
    df = df[~df['ProductID'].isin(form_df['ProductID'])]

    # Write the updated DataFrame back to the Excel file
    with pd.ExcelWriter('Excel/handover_data.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Return a success message
    return jsonify({'message': 'Data updated successfully.'})


@app.route('/approve_receive_request', methods=['POST'])
def approve_receive_request():
    form_data = request.json  # Assuming the form data is sent as JSON
    print("This is the approve_receive_request form data", form_data)
    
    last_dict = form_data[0]  # Get the last dictionary from the list

    # Assign the values of the last dictionary to the global variables
    from_person = last_dict.get('FromPerson')
    to_person = last_dict.get('ToPerson')
    from_project = last_dict.get('FromProject')
    to_project = last_dict.get('ToProject')

    # Optionally, you can print the values for verification
    print("From Person:", from_person)
    print("To Person:", to_person)
    print("From Project:", from_project)
    print("To Project:", to_project)


    # Get the EwayBill value from the first dictionary
    eway_bill_value = form_data[1].get('EwayBill', None)
    formid = form_data[2].get('FormNo', None)
    print("formid", formid)

    text = "Receive Approval Form"
    send_email(text, formid, eway_bill_value,from_person,to_person,from_project ,to_project)
    ''' 
    if form_data:


        # Open inventory.xlsx
        excel_file = "Excel/inventory.xlsx"
        df = pd.read_excel(excel_file)

        # Iterate over the rest of the dictionaries
        for item in form_data[4:]:
            product_id = item['ProductID']
            condition = item['Condition']

            # Update CurrentProject, CurrentOwner, and Condition columns for each product ID
            df.loc[df['ProductID'] == product_id, 'CurrentProject'] = current_project
            df.loc[df['ProductID'] == product_id, 'CurrentOwner'] = current_person
            df.loc[df['ProductID'] == product_id, 'Condition'] = condition

        # Save the updated DataFrame back to the Excel file
        df.to_excel(excel_file, index=False)
        print("Data updated successfully.")

    # Open handover_data.xlsx
    excel_file_handover = "Excel/handover_data.xlsx"
    df_handover = pd.read_excel(excel_file_handover)

    # Check if there is any form data
    if form_data:
        product_ids = [item['ProductID'] for item in form_data[4:]]  # Start extracting from the fourth dictionary onwards
        df_handover = df_handover[~df_handover['ProductID'].isin(product_ids)]  # Remove rows with matching product IDs

        # Save the updated DataFrame back to the Excel file
        df_handover.to_excel(excel_file_handover, index=False)
    '''
    return "Approval to receive items is successfully given.\nName and project transfer has successfully completed.\nThe email is sent.\nThe receiver may proceed to utilize the items.\nThe form will self-destruct, Congrats for the successful transfer."



@app.route('/cart_items')
def cart_items():
    # Load the Excel file
    wb = load_workbook('Excel/inventory.xlsx')
    sheet = wb.active

    # Define a list to store the dictionaries
    data = []

    # Define a list to store ProductIDs
    product_ids = []

    # Get the name from session data
    name = "Fahad"
    
    # Iterate over rows and append data to the list as dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the current owner matches the name from the session cookie
        if row[8] == name:
            # Append the ProductID to the list
            product_ids.append(row[5])
            
            item = {
                'Category': row[1],
                'ProductID': row[5],
                'Name': row[2],
                'Make': row[3],
                'Model': row[4],
                'FromPerson' : row[8]
            }
            data.append(item)

    # Call extract_data_from_excel function to get result
    result = extract_data_from_excel(product_ids)
    
    # Replace NaN values with "NAN" in data
    data = [{k: "NAN" if pd.isna(v) else v for k, v in item.items()} for item in data]

    # Replace NaN values with "NAN" in result
    result = [[col if pd.notna(col) else "NAN" for col in row] for row in result]

    # Create a new list containing both data and result
    combined_data = [data, result]
    #print("this is the combined data",combined_data)
    return jsonify(combined_data=combined_data)



global extracted_rows_data
extracted_rows_data = None

@app.route('/get_list', methods=['GET'])
def get_list():

    global extracted_rows_data

    # Retrieve the text sent in the query parameter
    text_data = request.args.get('text', '')
    excel = "Excel/handover_data.xlsx"

    if text_data == "Approve Sender Form":
        column_name = "FromProject"
        project = session.get('login_row_data', {}).get('Project', 'Unknown')
        print("this is the project name value that we are trying to search",project)
        extracted_rows_data, form_id_count = extract_rows_first_filter(excel, column_name, project)
        print("This is the extracted rows data from the first filter",extracted_rows_data)
        return str(form_id_count)

    elif text_data == "Receiver Form":
        column_name = "ToPerson"
        project = session.get('login_row_data', {}).get('Name', 'Unknown')
        print("this is the project name value that we are trying to search",project)
        extracted_rows_data, form_id_count = extract_rows_first_filter(excel, column_name, project)
        print("This is the extracted rows data from the first filter",extracted_rows_data)
        return str(form_id_count)

    elif text_data == "Approve Receiver Form":
        column_name = "ToProject"
        project = session.get('login_row_data', {}).get('Project', 'Unknown')
        print("this is the project name value that we are trying to search",project)
        extracted_rows_data, form_id_count = extract_rows_first_filter(excel, column_name, project)
        print("This is the extracted rows data from the first filter",extracted_rows_data)
        return str(form_id_count)

    else:
        # Handle case when text doesn't match any condition
        total_rows = "Unknown text"

    return str(total_rows)












@app.route('/get_username')
def get_username():

    name = session.get('login_row_data', {}).get('Name', 'Unknown')


    return jsonify({'username': name})
        # Get the name from session data







@app.route('/recieve_table', methods=['GET'])
def recieve_table():
    # Load the data from the Excel file into a pandas DataFrame
    df = pd.read_excel('Excel/handover_data.xlsx')

    #print(df.columns)
    # Get the project from session data
    name = "Umar"

    # Filter the DataFrame based on the parameters
    filtered_data = df[(df['ToPerson'] == name) & (df['AskReceiveApproval'] == "yes")]
    print
    # Convert the filtered data to JSON format
    json_data = filtered_data.to_json(orient='records')

    return json_data








@app.route('/approval_table', methods=['GET'])
def approval_table():

    #
    # Load the data from the Excel file into a pandas DataFrame
    df = pd.read_excel('Excel/handover_data.xlsx')

    # Get the project from session data
    project = "SOI ASSAM"

    # Filter the DataFrame based on the parameters
    source_df = df[(df['FromProject'] == project) & (df["ApprovalToSend"] == "yes")]
    destination_df = df[(df['FromProject'] == project)].copy()  # Make a copy to avoid SettingWithCopyWarning
    
    # Update columns in the DataFrames
    destination_df["ApprovalType"] = "Receive"
    source_df["ApprovalType"] = "Send"

    # Append destination_df to source_df
    source_df = pd.concat([source_df, destination_df])

    # Sort the DataFrame based on "FormID"
    source_df = source_df.sort_values("FormID")

    # Convert the filtered data to JSON format
    json_data = source_df.to_json(orient='records')

    return json_data





more_df = pd.DataFrame()  # Initialize less_df as an empty DataFrame

@app.route('/transaction_history', methods=['GET'])
def transaction_history():

    global more_df

    # Load the data from the Excel file into a pandas DataFrame
    df = pd.read_excel('Excel/handover_data.xlsx')
    print("this is the excel data",df)
    # Get the parameters name and project from session data
    #name = request.args.get('name')
    #project = request.args.get('project')

    name = "Umar"
    project = "SOI ASSAM"

    # Filter the DataFrame based on the parameters
    filtered_data = df[(df['FromProject'] == project) | (df['FromPerson'] == name) | (df['ToProject'] == project) | (df['ToPerson'] == name)]
    more_df = filtered_data
    filtered_data = filtered_data.dropna(subset=['FormID'])

    print("this is the less df",more_df)
    
    # Convert the filtered data to JSON format
    json_data = filtered_data.to_json(orient='records')

    return json_data


@app.route('/my_invent_dashboard')
def my_invent_dashboard():
    try:
        # Get the name from session data
        name = session.get('login_row_data', {}).get('Name', 'Unknown')
        print("this is the global employee name", name)
        # Read Excel file
        excel_data = pd.read_excel("Excel/inventory.xlsx")
        #print("this is the excel data", excel_data)
        # Replace NaN values with 'NAN'
        excel_data = excel_data.fillna('NAN')
        #print("this is the excel data", excel_data)
        # Filter data based on 'CurrentOwner' column
        filtered_data = excel_data[excel_data['CurrentOwner'] == name]
        #print("this is the filtered data",filtered_data)
        # Remove the last two columns
        filtered_data = filtered_data.iloc[:, :-1]

        # Convert data to list of dictionaries
        data_list = filtered_data.to_dict(orient='records')
        #print("this is the data listttttttttttttttttttttttttt",data_list)
        # Return data as JSON
        return jsonify(data_list)
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/invent_dashboard')
def invent_dashboard():
    try:
        # Read Excel file
        excel_data = pd.read_excel("Excel/inventory.xlsx")
        
        # Replace NaN values with 'NAN'
        excel_data = excel_data.fillna('nan')
        
        # Convert data to list of dictionaries
        data_list = excel_data.to_dict(orient='records')

        # Return data as JSON
        return jsonify(data_list)
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/my_project_dashboard')
def my_project_dashboard():
    try:
        # Read Excel file
        excel_data = pd.read_excel("Excel/inventory.xlsx")
        
        # Replace NaN values with 'NAN'
        excel_data = excel_data.fillna('nan')
        print('excel data',excel_data)
        # Filter rows where 'FromProject' or 'ToProject' column matches the project variable
        project = 'SOI Tripura'
        filtered_data = excel_data[(excel_data['CurrentProject'] == project)]
        
        # Convert filtered data to list of dictionaries
        data_list = filtered_data.to_dict(orient='records')

        # Return filtered data as JSON
        return jsonify(data_list)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/transferstatus')
def transferstatus():
    try:
        # Read the Excel file
        df = pd.read_excel('Excel/handover_data.xlsx')  # Update with your file path
        
        # Filter rows where FormID column is not null
        filtered_df = df[df['FormID'].notnull()]
        
        # Convert filtered data to JSON
        json_data = filtered_df.to_json(orient='records')
        
        return json_data
    except Exception as e:
        return jsonify({'error': str(e)}), 500



if __name__ == "__main__":
    app.run(debug=True, port=5001)